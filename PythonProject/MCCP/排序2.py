import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

def load_data(file_path):
    """加载并整理Excel数据"""
    sheets = pd.read_excel(file_path, sheet_name=None)
    all_data = []
    
    for sheet_name, df in sheets.items():
        # 统一处理不同格式的工作表
        if df.shape[1] >= 2:  # 至少有两列
            # 尝试自动识别教师和得分列
            teacher_col = next((col for col in df.columns if '教师' in str(col)), df.columns[0])
            score_col = next((col for col in df.columns if '得分' in str(col)), df.columns[1])
            
            temp = df[[teacher_col, score_col]].copy()
            temp.columns = ['教师', '得分']
            temp['学院'] = sheet_name
            all_data.append(temp.dropna())
    
    return pd.concat(all_data)

def process_and_sort(df):
    """处理数据并按学院升序排列"""
    # 按学院名称升序排序
    df_sorted = df.sort_values('学院', ascending=True)
    
    # 添加学院内排名
    df_sorted['学院内排名'] = df_sorted.groupby('学院')['得分'].rank(ascending=False, method='min')
    
    return df_sorted

def save_sorted_results(df, output_file):
    """将排序结果保存到Excel"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 按学院分组保存
        for college, group in df.groupby('学院'):
            # 每个学院一个工作表
            group.to_excel(writer, sheet_name=college[:31], index=False)  # 限制sheet名称长度
            
            # 获取工作表对象设置格式
            sheet = writer.sheets[college[:31]]
            
            # 设置标题样式
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
                sheet.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", fill_type="solid")
                sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
            
            # 设置列宽
            sheet.column_dimensions['A'].width = 15  # 教师列
            sheet.column_dimensions['B'].width = 10  # 得分列
            sheet.column_dimensions['C'].width = 20  # 学院列
            sheet.column_dimensions['D'].width = 12  # 排名列
        
        # 添加汇总表
        summary = df.groupby('学院')['得分'].agg(['count', 'mean', 'max', 'min']).reset_index()
        summary.columns = ['学院', '教师人数', '平均分', '最高分', '最低分']
        summary.to_excel(writer, sheet_name='学院汇总', index=False)
        
        # 设置汇总表格式
        sheet = writer.sheets['学院汇总']
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
            sheet.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", fill_type="solid")
        
        # 添加学院名称升序说明
        sheet.cell(row=sheet.max_row + 2, column=1, value="※ 按学院名称拼音升序排列")

# 使用示例
if __name__ == '__main__':
    input_file = '教师评分标准化结果.xlsx'
    output_file = '按学院排序结果.xlsx'
    
    print("正在加载数据...")
    raw_data = load_data(input_file)
    
    print("正在处理排序...")
    sorted_data = process_and_sort(raw_data)
    
    print("正在生成Excel文件...")
    save_sorted_results(sorted_data, output_file)
    
    print(f"处理完成！结果已保存到: {output_file}")
    print(f"共处理 {len(sorted_data)} 条记录，来自 {sorted_data['学院'].nunique()} 个学院")
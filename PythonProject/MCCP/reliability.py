import math
from math import e
import openpyxl as op

file_path_icc = 'D:\Codes\数模校赛代码\ICC_colum1.xlsx'
wb = op.load_workbook(file_path_icc)
ws1 = wb["Sheet1"]
icc_list = []
for i in range(2,21):
    cell_value = ws1[f"B{i}"].value
    #print(cell_value)
    icc_list.insert(i-2,cell_value)
#print(icc_list)

file_path_zsc = "D:\Codes\数模校赛代码\zscr fnl.xlsx"
wb2 = op.load_workbook(file_path_zsc)
ws2 = wb2["Sheet1"]
zsc_list = []
for i in range(2,21):
    cell_value2=ws2[f"B{i}"].value
    #print(cell_value2)
    zsc_list.insert(i-2,cell_value2)
#print(zsc_list)

Zamax = 4
Zamin = 2.5
Imin = 0.4
Imax = 0.6
file_path_final = r"D:\Codes\数模校赛代码\可信度.xlsx"
wb3 = op.load_workbook(file_path_final)
ws3 = wb3["Sheet1"]
for i in range(2,21) :
    icc_part = pow(e,(abs(icc_list[i-2]-Imin)/abs(Imax-Imin)))
    zsc_part = (abs(zsc_list[i-2]-Zamin)*Zamax)/abs(Zamax-Zamin)
    Crexx  = icc_part - zsc_part
    Cre = (3+Crexx)/(e+3)
    if i<11 :
        ws3.cell(row=i,column=2).value = Cre
        #print(f"专家{i-1}的可信度为",Cre)
    else :
        ws3.cell(row=i,column=2).value = Cre
        #print(f"专家{i}的可信度为",Cre)
wb3.save("可信度.xlsx")
wb3.close()
    

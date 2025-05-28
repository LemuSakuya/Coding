import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def load_and_clean_data(file_path):
    """智能加载并清洗Excel数据"""
    sheets = pd.read_excel(file_path, sheet_name=None)
    cleaned_data = []
    
    for sheet_name, df in sheets.items():
        # 跳过空表
        if df.empty:
            continue
            
        # 寻找包含教师和得分的列
        teacher_col = None
        score_col = None
        
        for col in df.columns:
            col_str = str(col).lower()
            if '教师' in col_str or 'teacher' in col_str or '姓名' in col_str:
                teacher_col = col
            elif '得分' in col_str or 'score' in col_str or '评分' in col_str:
                score_col = col
        
        # 如果找到所需列
        if teacher_col and score_col:
            temp = df[[teacher_col, score_col]].copy()
            temp.columns = ['教师', '得分']
            temp['学院'] = sheet_name
            cleaned_data.append(temp.dropna())
    
    return pd.concat(cleaned_data).reset_index(drop=True)

def calculate_scores(df):
    """计算标准化分数"""
    # 学院内Z-score标准化
    df['Z分数'] = df.groupby('学院')['得分'].transform(
        lambda x: (x - x.mean()) / x.std())
    
    # 学院内百分位排名
    df['百分位'] = df.groupby('学院')['得分'].transform(
        lambda x: x.rank(pct=True))
    
    # 综合评分（可调整权重）
    df['综合评分'] = 0.6*df['Z分数'] + 0.4*df['百分位']
    
    # 按综合评分排序
    df = df.sort_values('综合评分', ascending=False)
    return df

def save_to_excel(df, output_file):
    """将结果保存为格式化的Excel文件"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. 保存原始数据
        df[['学院', '教师', '得分']].to_excel(
            writer, sheet_name='原始数据', index=False)
        
        # 2. 保存标准化结果
        df.to_excel(writer, sheet_name='标准化结果', index=False)
        
        # 3. 学院统计信息
        stats = df.groupby('学院').agg({
            '得分': ['count', 'mean', 'std', 'min', 'median', 'max'],
            '综合评分': ['mean', 'std']
        })
        stats.to_excel(writer, sheet_name='学院统计')
        
        # 获取工作簿对象
        workbook = writer.book
        
        # 设置单元格样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal='center')
        
        # 应用样式到所有工作表
        for sheetname in writer.sheets:
            sheet = writer.sheets[sheetname]
            
            # 设置标题样式
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # 设置数据格式
            if sheetname == '标准化结果':
                for row in range(2, sheet.max_row + 1):
                    for col in range(3, sheet.max_column + 1):  # 从第3列开始是数值
                        sheet.cell(row=row, column=col).number_format = '0.000'
            
            # 自动调整列宽
            for column in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

# 主程序
if __name__ == '__main__':
    input_file = '附件2.xlsx'
    output_file = '教师评分标准化结果.xlsx'
    
    try:
        print("正在加载数据...")
        df = load_and_clean_data(input_file)
        
        print("正在计算标准化分数...")
        df = calculate_scores(df)
        
        print("正在生成Excel报告...")
        save_to_excel(df, output_file)
        
        print(f"处理完成！结果已保存到: {output_file}")
        print(f"共处理 {len(df)} 条教师记录，来自 {df['学院'].nunique()} 个学院")
        
    except Exception as e:
        print(f"处理失败: {str(e)}")